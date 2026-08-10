import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlmodel import Session, select
from app.config import OUTPUTS_DIR, INPUTS_DIR, BASE_DIR
from app.db.database import engine
from app.db.models import Transaction

# Primary User Vendor Mapping File
INPUT_VENDOR_MAPPING_PATH = BASE_DIR / "data" / "inputs" / "Vendor_Category_Mapping.xlsx"
DROPBOX_MAPPING_PATH = INPUTS_DIR / "Vendor_Category_Mapping.xlsx"
OUTPUT_MAPPING_PATH = OUTPUTS_DIR / "Vendor_Category_Mapping.xlsx"

def get_active_mapping_path():
    if INPUT_VENDOR_MAPPING_PATH.exists():
        return INPUT_VENDOR_MAPPING_PATH
    elif DROPBOX_MAPPING_PATH.exists():
        return DROPBOX_MAPPING_PATH
    return OUTPUT_MAPPING_PATH

def generate_vendor_category_file(overwrite: bool = False) -> str:
    """STRICT READ-ONLY PROTECTION: Never overwrite user Vendor_Category_Mapping.xlsx file."""
    target_path = get_active_mapping_path()
    if target_path.exists() and not overwrite:
        return str(target_path)
        
    with Session(engine) as session:
        txs = session.exec(select(Transaction)).all()
        
    vendor_stats = {}
    for t in txs:
        v = t.vendor.strip()
        if v not in vendor_stats:
            vendor_stats[v] = {
                "vendor": v,
                "current_category": t.category or "Uncategorized",
                "current_subcategory": t.subcategory or "",
                "count": 0,
                "total_spent_ils": 0.0,
                "cards": set(),
                "users": set(),
                "source_files": set(),
                "last_date": t.transaction_date
            }
        vendor_stats[v]["count"] += 1
        vendor_stats[v]["total_spent_ils"] += t.charged_amount
        if t.card_last_4 and t.card_last_4 != "0000":
            vendor_stats[v]["cards"].add(t.card_last_4)
        if t.user_name and t.user_name != "Unassigned":
            vendor_stats[v]["users"].add(t.user_name)
        if t.source_file:
            vendor_stats[v]["source_files"].add(t.source_file)
        if t.transaction_date > vendor_stats[v]["last_date"]:
            vendor_stats[v]["last_date"] = t.transaction_date

    sorted_vendors = sorted(vendor_stats.values(), key=lambda x: x["vendor"].lower())
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendor Categories"
    
    headers = [
        "Vendor Name",
        "Category",
        "Subcategory",
        "Transaction Count",
        "Total Spent (ILS)",
        "Cards / Accounts",
        "Family User(s)",
        "Source File(s)",
        "Latest Transaction Date"
    ]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for item in sorted_vendors:
        cards_str = ", ".join([f"Card {c}" for c in sorted(item["cards"])]) if item["cards"] else "N/A"
        users_str = ", ".join(sorted(item["users"])) if item["users"] else "Unassigned"
        files_str = ", ".join(sorted(item["source_files"])) if item["source_files"] else "N/A"
        
        ws.append([
            item["vendor"],
            item["current_category"],
            item["current_subcategory"],
            item["count"],
            item["total_spent_ils"],
            cards_str,
            users_str,
            files_str,
            item["last_date"].strftime("%Y-%m-%d")
        ])
        
    for row in ws.iter_rows(min_row=2):
        for idx, cell in enumerate(row):
            cell.font = data_font
            cell.border = thin_border
            if idx == 3:
                cell.alignment = Alignment(horizontal="center")
            elif idx == 4:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
                
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    wb.save(target_path)
    return str(target_path)

def apply_vendor_category_file() -> int:
    """STRICTLY READS Vendor_Category_Mapping.xlsx from data/inputs folder."""
    source_path = get_active_mapping_path()
    if not source_path.exists():
        return 0
        
    df = pd.read_excel(source_path)
    
    col_vendor = [c for c in df.columns if 'vendor' in str(c).lower() or 'בית עסק' in str(c).lower() or 'שם' in str(c).lower()][0]
    col_cat_candidates = [c for c in df.columns if 'category' in str(c).lower() and 'sub' not in str(c).lower() or 'קטגוריה' in str(c).lower() or 'ענף' in str(c).lower()]
    col_cat = col_cat_candidates[0] if col_cat_candidates else df.columns[1]
    
    col_subcat_candidates = [c for c in df.columns if 'sub' in str(c).lower() or 'תת' in str(c).lower()]
    col_subcat = col_subcat_candidates[0] if col_subcat_candidates else (df.columns[2] if len(df.columns) > 2 else None)

    mapping = {}
    for _, row in df.iterrows():
        vendor = str(row[col_vendor]).strip()
        cat = str(row[col_cat]).strip() if pd.notna(row[col_cat]) else "Uncategorized"
        subcat = str(row[col_subcat]).strip() if col_subcat and pd.notna(row[col_subcat]) else None
        mapping[vendor] = (cat if cat and cat.lower() not in ['nan', 'none'] else "Uncategorized", 
                           subcat if subcat and subcat.lower() not in ['nan', 'none'] else None)
            
    updated_count = 0
    with Session(engine) as session:
        txs = session.exec(select(Transaction)).all()
        for t in txs:
            v = t.vendor.strip()
            if v in mapping:
                target_cat, target_subcat = mapping[v]
                if t.category != target_cat or t.subcategory != target_subcat:
                    t.category = target_cat
                    t.subcategory = target_subcat
                    session.add(t)
                    updated_count += 1
        session.commit()
        
    return updated_count

if __name__ == "__main__":
    apply_vendor_category_file()
