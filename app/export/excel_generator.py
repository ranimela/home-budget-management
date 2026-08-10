import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlmodel import Session, select
from app.config import MASTER_EXCEL_PATH
from app.db.database import engine
from app.db.models import Transaction, CardMapping

CARD_DISPLAY_NAMES = {
    '9380': 'Rani CAL',
    '4591': 'Rani Leumicard',
    '4656': 'Yael Mastercard',
    '1123': 'Yael Max'
}

def generate_master_excel() -> str:
    with Session(engine) as session:
        txs = session.exec(select(Transaction)).all()
        card_mappings = {m.card_last_4: m.display_name for m in session.exec(select(CardMapping)).all()}

    rows = []
    for t in txs:
        card_label = CARD_DISPLAY_NAMES.get(t.card_last_4, card_mappings.get(t.card_last_4, f"Card {t.card_last_4}"))
        rows.append({
            "ID": t.id,
            "Transaction Date": t.transaction_date.strftime("%d/%m/%y"),
            "Actual Billing Date": t.charge_date.strftime("01/%m/%y") if t.charge_date else t.transaction_date.strftime("01/%m/%y"),
            "User": t.user_name or "Unassigned",
            "Card Name": card_label,
            "Card Last 4": t.card_last_4,
            "Institution": t.institution,
            "Vendor": t.vendor,
            "Category": t.category or "Uncategorized",
            "Subcategory": t.subcategory or "",
            "Charged Amount (ILS)": t.charged_amount,
            "Original Amount": t.original_amount,
            "Currency": t.original_currency,
            "Installment": f"{t.current_installment}/{t.total_installments}" if t.total_installments > 1 else "Single",
            "Source File": t.source_file
        })

    df_tx = pd.DataFrame(rows)

    wb = Workbook()
    
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    # 1. All Transactions Sheet
    ws_all = wb.active
    ws_all.title = "All Transactions"
    _write_styled_df(ws_all, df_tx, header_fill, header_font, data_font, thin_border)

    # 2. Monthly Summary Sheet
    ws_summary = wb.create_sheet(title="Monthly Summary")
    if not df_tx.empty:
        summary_df = df_tx.groupby(['Actual Billing Date', 'Category'])['Charged Amount (ILS)'].agg(['sum', 'count']).reset_index()
        summary_df.columns = ['Billing Month', 'Category', 'Total Spent (ILS)', 'Transaction Count']
        summary_pivot = summary_df.pivot(index='Category', columns='Billing Month', values='Total Spent (ILS)').fillna(0)
        summary_pivot.reset_index(inplace=True)
        _write_styled_df(ws_summary, summary_pivot, header_fill, header_font, data_font, thin_border)
    else:
        _write_styled_df(ws_summary, pd.DataFrame(), header_fill, header_font, data_font, thin_border)

    # 3. Card & User Breakdown Sheet
    ws_user = wb.create_sheet(title="Card & User Breakdown")
    if not df_tx.empty:
        user_df = df_tx.groupby(['User', 'Card Name'])['Charged Amount (ILS)'].agg(['sum', 'count']).reset_index()
        user_df.columns = ['Family User', 'Card Name', 'Total Spent (ILS)', 'Transaction Count']
        _write_styled_df(ws_user, user_df, header_fill, header_font, data_font, thin_border)
    else:
        _write_styled_df(ws_user, pd.DataFrame(), header_fill, header_font, data_font, thin_border)

    # 4. Installments Schedule Sheet
    ws_inst = wb.create_sheet(title="Installments Schedule")
    df_inst = df_tx[df_tx['Installment'] != "Single"] if not df_tx.empty else pd.DataFrame()
    _write_styled_df(ws_inst, df_inst, header_fill, header_font, data_font, thin_border)

    # 5. Budget & 12M Projections Sheet
    ws_proj = wb.create_sheet(title="Budget & 12M Projections")
    from app.core.projections import calculate_monthly_projections
    proj_matrix = calculate_monthly_projections(12)
    proj_data = []
    for p in proj_matrix:
        proj_data.append({
            "Month": p["month"],
            "Active Installments Count": p["active_installments_count"],
            "Installments Commitment (ILS)": p["projected_installments_ils"],
            "Baseline Category Average (ILS)": p["projected_variable_baseline_ils"],
            "Total Projected Monthly Spend (ILS)": p["total_projected_spend_ils"]
        })
    df_proj = pd.DataFrame(proj_data)
    _write_styled_df(ws_proj, df_proj, header_fill, header_font, data_font, thin_border)

    try:
        wb.save(MASTER_EXCEL_PATH)
    except PermissionError:
        alt_path = MASTER_EXCEL_PATH.parent / "Master_Budget_Updated.xlsx"
        wb.save(alt_path)
        return str(alt_path)
    return str(MASTER_EXCEL_PATH)

def _write_styled_df(ws, df, header_fill, header_font, data_font, border):
    if df.empty:
        ws.append(["No Data"])
        return

    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in df.iterrows():
        ws.append(list(row.values))

    for row in ws.iter_rows(min_row=2):
        for idx, cell in enumerate(row):
            cell.font = data_font
            cell.border = border
            val_str = str(cell.value)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif val_str.count('/') == 2 and len(val_str) == 8:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)
